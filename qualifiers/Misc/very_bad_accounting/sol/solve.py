from pathlib import Path
from openpyxl import load_workbook
import re

wb = load_workbook(Path(__file__).parent.parent / "dist" / "equilibrium.xlsx")
ws = wb["Ouroboros"]

pat = re.compile(r"^=F(\d+)/2\+(\d+)/2$")
targets = {
    cell.row: int(m.group(2))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row)
    for cell in [row[5]]
    if cell.value and isinstance(cell.value, str) and (m := pat.match(cell.value)) and int(m.group(1)) == cell.row
}

state = {r: 0.0 for r in targets}
for _ in range(100):
    state = {r: state[r] / 2 + t / 2 for r, t in targets.items()}

print("".join(chr(round(v)) for v in state.values()))
